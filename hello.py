from os.path import sep
import time
from openpyxl import load_workbook
from pandas_datareader import data
import fix_yahoo_finance as yf
yf.pdr_override()
# start_date = '1996-05-06' #startdate를 1996년으로 설정해두면 가장 오래된 데이터부터 전부 가져올 수 있다.
start_date = '1996-05-06'
tickers = ['067160.KQ', '035420.KS','035720.KS'] #1 아프리카tv와 네이버의 ticker(종목코드)
afreeca = data.get_data_yahoo(tickers[0], start_date)
naver = data.get_data_yahoo(tickers[1], start_date)
daum = data.get_data_yahoo(tickers[2], start_date)
naver.to_excel('hello.xlsx')
wb = load_workbook("hello.xlsx")
ws = wb.active
s = []
first = []
highest = 0
date = ""
for i in range(2,ws.max_row+1):
    f = i - 1
    if f == 1:
        f = 2
    a = ws.cell(row = f,column = 3).value
    b = ws.cell(row = i,column = 3).value
    first.append(b)
    m = b/a*100-100
    mh = int(m)
    s.append(mh)
for gh in range(2,ws.max_row+1):
    if gh + 30 > ws.max_row:
        ff = ws.max_row
    else:
        ff = gh + 30
    ab = ws.cell(row = ff,column = 4).value
    bc = ws.cell(row = gh,column = 3).value
    rs = int(ab - bc)
    if rs > highest:
        print(highest,date,sep=" ")
        highest = rs
        date = ws.cell(row = gh,column = 1).value
print(highest,date,sep=" ")

